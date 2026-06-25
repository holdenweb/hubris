from pathlib import Path

import pytest

from hubris import read_file

FIXTURE = Path(__file__).parent / "data" / "parameters.xlsx"


@pytest.fixture(scope="module")
def params():
    return read_file(FIXTURE)


def test_scalar(params):
    assert params["title"] == "Hubris Demo"


def test_named_range_dict_hours(params):
    assert params["hours"]["Monday"] == "7:00 am - 8:00 pm"
    assert params["hours"]["Sunday"] == "Closed"
    assert len(params["hours"]) == 7


def test_named_range_dict_prices(params):
    assert params["prices"] == {"Tea": 3.25, "Coffee": 4.0, "Bacon Sandwich": 8.25}


def test_matrix_rows(params):
    assert params["hours_list"][0] == ["Monday", "7:00 am - 8:00 pm"]
    assert params["hours_list"][5] == ["Saturday", "9:00 am - 5:00 pm"]
    assert len(params["hours_list"]) == 7


def test_horizontal_vector(params):
    assert params["h_vector"] == ["the", "quick", "brown", "fox"]


def test_vertical_vector(params):
    assert params["v_vector"] == ["jumps", "over", "the", "lazy", "dog"]


def test_expected_keys_present(params):
    assert set(params) >= {"title", "hours", "prices", "hours_list", "h_vector", "v_vector"}


def test_single_entry_dict(tmp_path):
    """A one-row {dict} range must read as a dict, not be flattened to a vector."""
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"], ws["B1"] = "version", "{version}"
    ws["A2"], ws["B2"] = "number", "0.1.2"
    wb.defined_names.add(DefinedName("Parameters", attr_text="Sheet1!$A$1:$B$1"))
    wb.defined_names.add(DefinedName("version", attr_text="Sheet1!$A$2:$B$2"))
    path = tmp_path / "single.xlsx"
    wb.save(path)

    assert read_file(path) == {"version": {"number": "0.1.2"}}
