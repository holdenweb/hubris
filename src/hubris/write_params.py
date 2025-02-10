from itertools import count

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

def link_num():
    yield from count()

class Allocator:

    def __init__(self):
        self.sheet = sheet
        self.next_col = 1

    def allocate(self, n):
        result = self.next_col
        self.next_col += n
        return result

link_num = count()
allocator = Allocator()

def dict_to_spreadsheet(data, filename="output.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active  # NO!! Find some better way to locate the extraction data
    sheet.title = "Parameters"  # Set the sheet title to "Parameters"

    def dict_to_sheet:





    def dict_to_sheet(data, start_row=1, start_col=1):
        row = start_row
        for key, value in data.items():
            sheet.cell(row=row, column=start_col).value = key

            if isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        link_text = f"range_{next(link_num)}"
                        sheet.cell(row=row, column=start_col + 1).value = link_text
                        row += 1
                        # Recursively process the nested dictionary
                        row = dict_to_sheet(item, row, start_col + 2)

                        # Define the named range for the nested dictionary
                        range_start_cell = openpyxl.utils.cell.get_column_letter(start_col + 1) + str(start_row)
                        range_end_cell = openpyxl.utils.cell.get_column_letter(start_col + 1) + str(row - 1)
                        range_string = f'{sheet.title}!{range_start_cell}:{range_end_cell}'
                        defined_name = DefinedName(name=link_text, attr_text=range_string)
                        workbook.defined_names.add(defined_name)

                    else:
                        sheet.cell(row=row, column=start_col + 1).value = item
                        row += 1
            elif isinstance(value, dict):
                sheet.cell(row=row, column=start_col + 1).value = f"[{key}_range]"
                row += 1
                # Recursively process the nested dictionary
                row = dict_to_sheet(value, row, start_col + 2)

                # Define the named range for the nested dictionary
                range_start_cell = openpyxl.utils.cell.get_column_letter(start_col + 1) + str(start_row)
                range_end_cell = openpyxl.utils.cell.get_column_letter(start_col + 1) + str(row - 1)
                range_string = f"'{sheet.title}'!{range_start_cell}:{range_end_cell}"
                defined_name = DefinedName(name=f"{key}_range", attr_text=range_string)
                workbook.defined_names.add(defined_name)
            else:
                sheet.cell(row=row, column=start_col + 1).value = value
                row += 1

        return row

    last_row = dict_to_sheet(data)

    # Define the "Parameters" named range
    range_start_cell = openpyxl.utils.cell.get_column_letter(1) + str(1)
    range_end_cell = openpyxl.utils.cell.get_column_letter(2) + str(last_row - 1)
    range_string = f"{sheet.title}!{range_start_cell}:{range_end_cell}"

    parameters_range = DefinedName(name="Parameters", attr_text=range_string)
    workbook.defined_names.add(parameters_range)

    # Save the workbook
    workbook.save(filename)